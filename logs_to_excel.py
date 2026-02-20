"""Дублирование логов в локальный Excel-файл (logs/logs.xlsx).
Каждое событие (Normalization, Judge, Feedback, Escalation) дописывается в соответствующий лист.
Запись в фоновом потоке, чтобы не блокировать бота. Требуется openpyxl."""
import logging
import os
import threading
from typing import Dict, Any, List

logger = logging.getLogger(__name__)

_lock = threading.Lock()
_excel_path = None


def _get_excel_path() -> str:
    global _excel_path
    if _excel_path is None:
        import config
        log_dir = os.path.abspath(getattr(config, "LOGS_PATH", "./logs"))
        os.makedirs(log_dir, exist_ok=True)
        _excel_path = os.path.join(log_dir, "logs.xlsx")
    return _excel_path


def _ensure_sheet(wb, sheet_name: str, headers: List[str]):
    """Создаёт лист с заголовками, если его нет или он пустой."""
    if sheet_name not in wb.sheetnames:
        ws = wb.create_sheet(sheet_name)
        ws.append(headers)
        return ws
    ws = wb[sheet_name]
    if ws.max_row == 0:
        ws.append(headers)
    return ws


def _append_normalization_sync(entry: Dict[str, Any]) -> None:
    try:
        import openpyxl
    except ImportError:
        logger.debug("openpyxl не установлен — запись в Excel отключена")
        return
    path = _get_excel_path()
    headers = ["timestamp", "user_id", "original_text", "normalized_query", "type"]
    with _lock:
        try:
            if os.path.exists(path):
                wb = openpyxl.load_workbook(path)
            else:
                wb = openpyxl.Workbook()
                wb.remove(wb.active)
            ws = _ensure_sheet(wb, "Normalization", headers)
            ws.append([
                entry.get("timestamp", ""),
                entry.get("user_id", ""),
                (str(entry.get("original_text") or ""))[:1000],
                (str(entry.get("normalized_query") or ""))[:500],
                entry.get("type", ""),
            ])
            wb.save(path)
        except Exception as e:
            logger.warning("Excel append Normalization: %s", e)


def _append_judge_sync(entry: Dict[str, Any]) -> None:
    try:
        import openpyxl
    except ImportError:
        return
    path = _get_excel_path()
    headers = [
        "timestamp", "request_id", "user_id", "question", "answer",
        "rel", "grnd", "safe", "compl",
        "verdict", "score", "type_ok", "refusal_ok", "explanation",
    ]
    j = entry.get("judge_verdict") or {}
    row = [
        entry.get("timestamp", ""),
        entry.get("request_id", ""),
        entry.get("user_id", ""),
        (str(entry.get("question") or ""))[:500],
        (str(entry.get("answer") or ""))[:500],
        j.get("relevance", ""),
        j.get("groundedness", ""),
        j.get("safety", ""),
        j.get("completeness", ""),
        j.get("verdict", ""),
        j.get("overall_score", ""),
        j.get("question_type_correct", ""),
        j.get("correct_refusal", ""),
        (str(j.get("explanation") or ""))[:300],
    ]
    with _lock:
        try:
            if os.path.exists(path):
                wb = openpyxl.load_workbook(path)
            else:
                wb = openpyxl.Workbook()
                wb.remove(wb.active)
            ws = _ensure_sheet(wb, "Judge", headers)
            ws.append(row)
            wb.save(path)
        except Exception as e:
            logger.warning("Excel append Judge: %s", e)


def _append_feedback_sync(entry: Dict[str, Any]) -> None:
    try:
        import openpyxl
    except ImportError:
        return
    path = _get_excel_path()
    headers = ["timestamp", "request_id", "user_id", "query_type", "rating", "feedback_at", "question", "answer"]
    row = [
        entry.get("timestamp", ""),
        entry.get("request_id", ""),
        entry.get("user_id", ""),
        entry.get("query_type", ""),
        entry.get("rating") or "",
        entry.get("feedback_at", ""),
        (str(entry.get("question") or ""))[:500],
        (str(entry.get("answer") or ""))[:500],
    ]
    with _lock:
        try:
            if os.path.exists(path):
                wb = openpyxl.load_workbook(path)
            else:
                wb = openpyxl.Workbook()
                wb.remove(wb.active)
            ws = _ensure_sheet(wb, "Feedback", headers)
            ws.append(row)
            wb.save(path)
        except Exception as e:
            logger.warning("Excel append Feedback: %s", e)


def _update_feedback_rating_sync(request_id: str, rating: str, feedback_at: str) -> None:
    try:
        import openpyxl
    except ImportError:
        return
    path = _get_excel_path()
    if not os.path.exists(path):
        return
    with _lock:
        try:
            wb = openpyxl.load_workbook(path)
            if "Feedback" not in wb.sheetnames:
                return
            ws = wb["Feedback"]
            for row in range(2, ws.max_row + 1):
                if str(ws.cell(row=row, column=2).value) == str(request_id):
                    ws.cell(row=row, column=5, value=rating)
                    ws.cell(row=row, column=6, value=feedback_at)
                    break
            wb.save(path)
        except Exception as e:
            logger.warning("Excel update Feedback rating: %s", e)


def _append_escalation_sync(entry: Dict[str, Any]) -> None:
    try:
        import openpyxl
    except ImportError:
        return
    path = _get_excel_path()
    headers = ["timestamp", "user_id", "question", "answer", "escalated"]
    row = [
        entry.get("timestamp", ""),
        entry.get("user_id", ""),
        (str(entry.get("question") or ""))[:500],
        (str(entry.get("answer") or ""))[:500],
        "1",
    ]
    with _lock:
        try:
            if os.path.exists(path):
                wb = openpyxl.load_workbook(path)
            else:
                wb = openpyxl.Workbook()
                wb.remove(wb.active)
            ws = _ensure_sheet(wb, "Escalation", headers)
            ws.append(row)
            wb.save(path)
        except Exception as e:
            logger.warning("Excel append Escalation: %s", e)


def duplicate_normalization_to_excel(entry: Dict[str, Any]) -> None:
    threading.Thread(target=_append_normalization_sync, args=(entry,), daemon=True).start()


def duplicate_judge_to_excel(entry: Dict[str, Any]) -> None:
    threading.Thread(target=_append_judge_sync, args=(entry,), daemon=True).start()


def duplicate_feedback_to_excel(entry: Dict[str, Any]) -> None:
    threading.Thread(target=_append_feedback_sync, args=(entry,), daemon=True).start()


def duplicate_feedback_rating_update_to_excel(request_id: str, rating: str, feedback_at: str) -> None:
    threading.Thread(target=_update_feedback_rating_sync, args=(request_id, rating, feedback_at), daemon=True).start()


def duplicate_escalation_to_excel(entry: Dict[str, Any]) -> None:
    threading.Thread(target=_append_escalation_sync, args=(entry,), daemon=True).start()
